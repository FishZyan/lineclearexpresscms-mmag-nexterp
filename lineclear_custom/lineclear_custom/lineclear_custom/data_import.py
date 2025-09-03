import sys
import requests
import json
from datetime import datetime, timedelta
import csv
import frappe
import os
import time

"""bench console
    from lineclear_custom.lineclear_custom.data_import import import_data
    import_data("Sales Invoice", filename)
"""

company = "Line Clear Express Sdn Bhd"
currency = "MYR"
headers = {
        "Authorization": "Token 458ffead2fb02c7:3017acd39d84830",
        "Content-Type": "application/json"
    }

#0  1                2           3                 4                 5           6       7               8           9            10      11         12
#ID	Posting Date	Debtor Code	 Reference No	Payment Due Date	Subtotal	Tax	    Income Account	Lhdn Tax	Description	  Agent   Tax Rate   Project No
def import_sales_invoice(rows, writer):
    erp_url = "http://localhost/api/resource/Sales%20Invoice"
    
    for sales_invoice in rows:
        if (not sales_invoice[0]):
            writer.writerow(["No invoice number"])
            continue
        if frappe.db.exists("Sales Invoice", sales_invoice[0]):
            writer.writerow(["Existed", sales_invoice[0]])
            continue

        customer = get_customer(sales_invoice[2])
        if not customer:
            writer.writerow(["Customer Not Found", sales_invoice[0]])
            continue
        data = {
            "name" : sales_invoice[0],
            "company" : company,
            "posting_date": datetime.strptime(sales_invoice[1], "%d/%m/%Y").strftime("%Y-%m-%d"),
            # "posting_date": sales_invoice[1].strftime("%Y-%m-%d") if sales_invoice[1] else '',
            "set_posting_time": 1,
            "currency": currency,
            "selling_price_list": "Standard Selling",
            "custom_debtor_code": sales_invoice[2],
            "reference_no" : sales_invoice[3],
            "customer": customer,
            "due_date": datetime.strptime(sales_invoice[4], "%d/%m/%Y").strftime("%Y-%m-%d") if sales_invoice[4] else (datetime.strptime(sales_invoice[1], "%d/%m/%Y").strftime("%Y-%m-%d") + timedelta(days=14).strftime("%Y-%m-%d")),
            # "due_date": sales_invoice[4].strftime("%Y-%m-%d") if sales_invoice[4] else (sales_invoice[1].strftime("%Y-%m-%d") + timedelta(days=14).strftime("%Y-%m-%d")),
            "items": [{
                "item_code": "Courier Service",
                "qty": 1,
                "description": sales_invoice[9],
                "rate": float(sales_invoice[5]),
                # "income_account": sales_invoice[7],
                "income_account": "501-1006 - INCOME - COURIER (OMS) - LCESB",
                # "custom_tax_type": sales_invoice[8]
                "custom_tax_type": "02"
            }],
            # "custom_tax_type": sales_invoice[8],
            "custom_tax_type": "02",
            "taxes": [{
                "charge_type": "Actual",
                "tax_amount": float(sales_invoice[6]),
                "account_head": "GST - LCESB",
                "description": "Tax Amount"
            }],
            "agent": sales_invoice[10],
            "custom_tax_rate": sales_invoice[11] or '',
            "docstatus": 1,
            "disable_rounded_total": 1
        }

        response = requests.post(erp_url, headers=headers, data=json.dumps(data))
        if response.status_code == 200:
            continue
            # writer.writerow(["Success",  sales_invoice[0]])
        else:
            writer.writerow(["Error",  sales_invoice[0]])

def import_payment_entry(rows, writer):
    erp_url = "http://localhost/api/resource/Payment%20Entry"
    previous_id = ""
    first_row = True
    data = {}
    #0  1              2            3               4                   5                6              7        8               9                10                11            12
    #ID	Posting Date   Debtor Code	Payment Method	Account Paid From	Account Paid To  Total Amount	Remark	 Reference Type	 Reference Name	 Allocated Amount	Cheque No	Reference Date
    for payment_entry in rows:
        if(not payment_entry[0]): #if multiple references to the same payment entry
            if(previous_id == ""):
                continue

            try: #Check if sales invoice reference exists
                if(payment_entry[8] == "Sales Invoice"):
                    check_sales_invoice = frappe.get_doc("Sales Invoice", payment_entry[9])
                elif(payment_entry[8] == "Journal Entry"):
                    check_sales_invoice = frappe.get_doc("Journal Entry", payment_entry[9])
            except:
                check_sales_invoice = None
            if not check_sales_invoice:
                writer.writerow(["Reference Not Found",  payment_entry[0]])
                continue

            #check if the allocated amount is larger than the outstanding amount
            try:
                if(payment_entry[8] == "Sales Invoice"):
                    if(payment_entry[10] and round(float(payment_entry[10]),2) > round(float(check_sales_invoice.outstanding_amount),2)):
                        status = "Allocated amount is larger than outstanding amount for reference {}".format(payment_entry[9])
                        writer.writerow([status, payment_entry[0]])
                        continue
            except:
                writer.writerow(["Unexpected error", payment_entry[0]])
                continue
            #Assign the amount to the reference document
            if(not payment_entry[10]):
                if(payment_entry[8] == "Sales Invoice"):
                    check_sales_invoice = frappe.get_doc("Sales Invoice", payment_entry[9])
                    reference_item = {
                        "reference_doctype": payment_entry[8],
                        "reference_name": payment_entry[9],
                        "allocated_amount": round(float(check_sales_invoice.outstanding_amount), 2)
                    }
                else:
                    status = "Reference Type {} not supported currently".format(payment_entry[9])
                    writer.writerow([status, payment_entry[0]])
                    continue
                amount_submitted += round(float(check_sales_invoice.outstanding_amount), 2)
            else:
                reference_item = {
                    "reference_doctype": payment_entry[8],
                    "reference_name": payment_entry[9],
                    "allocated_amount": round(float(payment_entry[10]),2)
                }
                amount_submitted += round(float(payment_entry[10]), 2)
            
            #Check if total allocated amount is larger than total amount
            if amount_submitted > total_amount:
                writer.writerow(["Total allocated amount is larger than total amount", payment_entry[0]])
                continue

            reference.append(reference_item)
            continue
        else:
            if(not first_row):
                data["references"] = reference
                response = requests.post(erp_url, headers=headers, data=json.dumps(data))
                if response.status_code == 200:
                    continue
                    # writer.writerow(["Success", previous_id])
                else:
                    writer.writerow(["Error", previous_id])
            try:
                check = frappe.get_doc("Payment Entry", payment_entry[0])
                previous_id = ""
            except:
                check = None
            if check:
                writer.writerow(["Existed", payment_entry[0]])
                continue
            
            first_row = False
            previous_id = payment_entry[0]
            reference = []
            amount_submitted = 0.00
            total_amount = round(float(payment_entry[6]), 2)

            try: #Check if sales invoice reference exists
                if(payment_entry[8] == "Sales Invoice"):
                    check_sales_invoice = frappe.get_doc("Sales Invoice", payment_entry[9])
                elif(payment_entry[8] == "Journal Entry"):
                    check_sales_invoice = frappe.get_doc("Journal Entry", payment_entry[9])
            except:
                check_sales_invoice = None
            if not check_sales_invoice:
                writer.writerow(["Reference Not Found", payment_entry[0]])
                continue

            #check if the allocated amount is larger than the outstanding amount
            try:
                if(payment_entry[8] == "Sales Invoice"):
                    if(payment_entry[10] and round(float(payment_entry[10]),2) > round(float(check_sales_invoice.outstanding_amount),2)):
                        status = "Allocated amount is larger than outstanding amount for reference {}".format(payment_entry[9])
                        writer.writerow([status, payment_entry[0]])
                        continue
            except:
                writer.writerow(["Unexpected Error", payment_entry[0]])
                continue

            #Assign the amount to the reference document
            if(not payment_entry[10]):
                if(payment_entry[8] == "Sales Invoice"):
                    check_sales_invoice = frappe.get_doc("Sales Invoice", payment_entry[9])
                    reference_item = {
                        "reference_doctype": payment_entry[8],
                        "reference_name": payment_entry[9],
                        "allocated_amount": round(float(check_sales_invoice.outstanding_amount), 2)
                    } 
                else:
                    status = "Reference Type {} not supported currently".format(payment_entry[9])
                    writer.writerow([status, payment_entry[0]])
                    continue
                amount_submitted += round(float(check_sales_invoice.outstanding_amount), 2)
            else:
                reference_item = {
                    "reference_doctype": payment_entry[8],
                    "reference_name": payment_entry[9],
                    "allocated_amount": round(float(payment_entry[10]),2)
                }
                amount_submitted += round(float(payment_entry[10]), 2)
            
            #Check if total allocated amount is larger than total amount
            
            if amount_submitted > total_amount:
                writer.writerow(["Total allocated amount is larger than total amount", payment_entry[0]])
                continue

            reference.append(reference_item) 
            customer = get_customer(payment_entry[2])
            if not customer:
                writer.writerow(["Customer Not Found", payment_entry[0]])
                continue
            data = {
                "name" : payment_entry[0],
                "payment_type": "Receive",
                "company" : company,
                "party_type" : "Customer",
                "posting_date": datetime.strptime(payment_entry[1], "%d/%m/%Y").strftime("%Y-%m-%d"),
                "debtor_code": payment_entry[2],
                "party": customer,
                "party_name": customer,
                "mode_of_payment": payment_entry[3] or None,
                "paid_from": payment_entry[4],
                "paid_to": payment_entry[5],
                "paid_from_account_currency": currency,
                "paid_to_account_currency": currency,
                "remark": payment_entry[7],
                "paid_amount": round(float(payment_entry[6]), 2),
                "received_amount": round(float(payment_entry[6]), 2),
                "reference_no": payment_entry[11] or None,
                "reference_date": datetime.strptime(payment_entry[12], "%d/%m/%Y").strftime("%Y-%m-%d"),
                "docstatus": 1,
                "cost_center": "Main - LCESB"
            }
            continue

    if(len(data) != 0):
        data["references"] = reference
        
        response = requests.post(erp_url, headers=headers, data=json.dumps(data))
        if response.status_code == 200:
            pass
            # writer.writerow(["Success", payment_entry[0]])1
        else:
            writer.writerow(["Error", payment_entry[0]])

def import_credit_note(rows, writer):
    erp_url = "http://localhost/api/resource/Journal%20Entry"

    for credit_note in rows:
        try:
            check = frappe.get_doc("Journal Entry", credit_note[0])
        except:
            check = None
        if check:
            writer.writerow(["Existed", credit_note[0]])
            continue

        customer = get_customer(credit_note[2])
        if not customer:
            writer.writerow(["Customer Not Found", credit_note[0]])
            continue
        
        if(credit_note[11] != ""):
            try: #Check if sales invoice reference exists
                if(credit_note[11] == "Sales Invoice"):
                    check_sales_invoice = frappe.get_doc("Sales Invoice", credit_note[12])
                elif(credit_note[11] == "Journal Entry"):
                    check_sales_invoice = frappe.get_doc("Journal Entry", credit_note[12])
            except:
                check_sales_invoice = None
            if not check_sales_invoice:
                writer.writerow(["Reference Not Found",  credit_note[0]])
                continue

            credit_items = {
                "account": credit_note[8],
                "party_type": "Customer",
                "party": customer,
                "custom_description": credit_note[10],
                "debit_in_account_currency": 0,
                "credit_in_account_currency": round(float(credit_note[5]), 2),
                "reference_type": credit_note[11],
                "reference_name": credit_note[12]
            }
        else:
            credit_items = {
                "account": credit_note[8],
                "party_type": "Customer",
                "party": customer,
                "custom_description": credit_note[10],
                "debit_in_account_currency": 0,
                "credit_in_account_currency": round(float(credit_note[5]), 2)
            }
        debit_items = {
            "account": credit_note[7],
            "party_type": None,
            "party": None,
            "custom_description": None,
            "debit_in_account_currency": round((float(credit_note[5]) - float(credit_note[6])), 2),
            "credit_in_account_currency": 0,
            "reference_type": None,
            "reference_name": None
        }

        if(float(credit_note[6]) > 0):
            tax_item = {
                "account": credit_note[9],
                "party_type": None,
                "party": None,
                "custom_description": None,
                "debit_in_account_currency": round(float(credit_note[6]), 2),
                "credit_in_account_currency": 0,
                "reference_type": None,
                "reference_name": None
            }
            all_items = [credit_items, debit_items, tax_item]
        else:
            all_items = [credit_items, debit_items]
        
        #0   1              2           3           4       5           6           7               8               9            10           11              12               13                        14               15             16
        #ID	Posting Date	Debtor Code	User Remark	Agent	Net Total	Total Tax	Debit Account	Credit Account	Tax Account	 Description  Reference Type  Reference Name   Item Classification Code  LHDN Tax Type    Created By     Tax Rate
        if(float(credit_note[6]) > 0):
            data = {
                "name" : credit_note[0],
                "voucher_type": "Credit Note",
                "company" : company,
                "custom_created_by": "System",
                "posting_date": datetime.strptime(credit_note[1], "%d/%m/%Y").strftime("%Y-%m-%d"),
                "currency": currency,
                "customer": customer,
                "debtor_code" : credit_note[2],
                "user_remark": credit_note[3],
                "agent": credit_note[4],
                "net_total": round(float(credit_note[5]), 2),
                "custom_total_tax_amount": round(float(credit_note[6]), 2),
                # "custom_tax_rate": credit_note[16] or '',
                "accounts": all_items,
                # "item_classification_code": credit_note[13],
                "item_classification_code": "004",
                # "lhdn_tax_type": credit_note[14],
                "lhdn_tax_type": "02",
                "custom_created_by": credit_note[15],
                "taxes": [{
                    "charge_type": "Actual",
                    "tax_amount": round(float(credit_note[6]), 2),
                    "account_head": credit_note[9],
                    "description": "Tax Amount"
                }],
                "docstatus": 1
            }
        else:
            data = {
                "name" : credit_note[0],
                "voucher_type": "Credit Note",
                "company" : company,
                "custom_created_by": "System",
                "posting_date": datetime.strptime(credit_note[1], "%d/%m/%Y").strftime("%Y-%m-%d"),
                "currency": currency,
                "customer": customer,
                "debtor_code" : credit_note[2],
                "user_remark": credit_note[3],
                "agent": credit_note[4],
                "net_total": round(float(credit_note[5]), 2),
                "custom_total_tax_amount": round(float(credit_note[6]), 2),
                # "custom_tax_rate": credit_note[16] or '',
                "accounts": all_items,
                # "item_classification_code": credit_note[13],
                "item_classification_code": "004",
                # "lhdn_tax_type": credit_note[14],
                "lhdn_tax_type": "02",
                "custom_created_by": credit_note[15],
                "docstatus": 1
            }

        response = requests.post(erp_url, headers=headers, data=json.dumps(data))
        if response.status_code == 200:
            pass
            # writer.writerow(["Success", credit_note[0]])
        else:
            print(response.text)
            writer.writerow(["Error", credit_note[0]])

def import_debit_note(rows, writer):
    erp_url = "http://localhost/api/resource/Journal%20Entry"
    
    for debit_note in rows:
        try:
            check = frappe.get_doc("Journal Entry", debit_note[0])
        except:
            check = None
        if check:
            writer.writerow(["Existed", debit_note[0]])
            continue

        if(float(debit_note[5]) <= 0):
            writer.writerow(["Net total cannot be less than 0", debit_note[0]])
            continue

        customer = get_customer(debit_note[2])
        if not customer:
            writer.writerow(["Customer Not Found", debit_note[0]])
            continue
        
        if(debit_note[11] != ""):
            try: #Check if sales invoice reference exists
                if(debit_note[11] == "Sales Invoice"):
                    check_sales_invoice = frappe.get_doc("Sales Invoice", debit_note[12])
                elif(debit_note[11] == "Journal Entry"):
                    check_sales_invoice = frappe.get_doc("Journal Entry", debit_note[12])
            except:
                check_sales_invoice = None
            if not check_sales_invoice:
                writer.writerow(["Reference Not Found",  debit_note[0]])
                continue

            debit_items = {
                "account": debit_note[7],
                "party_type": "Customer",
                "party": customer,
                "custom_description": debit_note[10],
                "debit_in_account_currency": round(float(debit_note[5]), 2),
                "credit_in_account_currency": 0,
                "reference_type": debit_note[11],
                "reference_name": debit_note[12]
            }
        else:
            debit_items = {
                "account": debit_note[7],
                "party_type": "Customer",
                "party": customer,
                "custom_description": debit_note[10],
                "debit_in_account_currency": round(float(debit_note[5]), 2),
                "credit_in_account_currency": 0
            }
        credit_items = {
            "account": debit_note[8],
            "party_type": None,
            "party": None,
            "custom_description": None,
            "debit_in_account_currency": 0,
            "credit_in_account_currency": round((float(debit_note[5]) - float(debit_note[6])), 2),
            "reference_type": None,
            "reference_name": None
        }

        if(float(debit_note[6]) > 0):
            tax_item = {
                "account": debit_note[9],
                "party_type": None,
                "party": None,
                "custom_description": None,
                "debit_in_account_currency": 0,
                "credit_in_account_currency": round(float(debit_note[6]), 2),
                "reference_type": None,
                "reference_name": None
            }
            all_items = [credit_items, debit_items, tax_item]
        else:
            all_items = [credit_items, debit_items]
        
        #0   1              2           3           4       5           6           7               8               9            10           11              12               13                        14
        #ID	Posting Date	Debtor Code	User Remark	Agent	Net Total	Total Tax	Debit Account	Credit Account	Tax Account	 Description  Reference Type  Reference Name   Item Classification Code  LHDN Tax Type
        if(float(debit_note[6]) > 0):
            data = {
                "name" : debit_note[0],
                "voucher_type": "Debit Note",
                "company" : company,
                "custom_created_by": "System",
                "posting_date": datetime.strptime(debit_note[1], "%d/%m/%Y").strftime("%Y-%m-%d"),
                "currency": currency,
                "customer": customer,
                "debtor_code" : debit_note[2],
                "user_remark": debit_note[3],
                "agent": debit_note[4],
                "net_total": round(float(debit_note[5]), 2),
                "custom_total_tax_amount": round(float(debit_note[6]), 2),
                # "custom_tax_rate": debit_note[16] or '',
                "accounts": all_items,
                # "item_classification_code": debit_note[13],
                "item_classification_code": "004",
                # "lhdn_tax_type": debit_note[14],
                "lhdn_tax_type": "02",
                "custom_created_by": debit_note[15],
                "taxes": [{
                    "charge_type": "Actual",
                    "tax_amount": round(float(debit_note[6]), 2),
                    "account_head": debit_note[9],
                    "description": "Tax Amount"
                }],
                "docstatus": 1
            }
        else:
            data = {
                "name" : debit_note[0],
                "voucher_type": "Debit Note",
                "company" : company,
                "custom_created_by": "System",
                "posting_date": datetime.strptime(debit_note[1], "%d/%m/%Y").strftime("%Y-%m-%d"),
                "currency": currency,
                "customer": customer,
                "debtor_code" : debit_note[2],
                "user_remark": debit_note[3],
                "agent": debit_note[4],
                "net_total": round(float(debit_note[5]), 2),
                "custom_total_tax_amount": round(float(debit_note[6]), 2),
                # "custom_tax_rate": debit_note[16] or '',
                "accounts": all_items,
                # "item_classification_code": debit_note[13],
                "item_classification_code": "004",
                # "lhdn_tax_type": debit_note[14],
                "lhdn_tax_type": "02",
                "custom_created_by": debit_note[15],
                "docstatus": 1
            }

        response = requests.post(erp_url, headers=headers, data=json.dumps(data))
        if response.status_code == 200:
            continue
            # writer.writerow(["Success", debit_note[0]])
        else:
            writer.writerow(["Error", debit_note[0]])

def get_customer(debtor_code):
    try:
        customer = frappe.get_doc("Customer", {"debtor_code": debtor_code})
        return customer.name
    except frappe.DoesNotExistError:
        customer = None
        return None
    
def get_unique_filename(filename):
    #status file name sales_invoice_status-YYMM-###.csv
    # Get current year and month in YYMM format
    date_part = datetime.now().strftime("%y%m")
    base_filename = f"{filename}-{date_part}"
    ext = ".csv"
    directory = frappe.get_site_path("private", "files")
    os.makedirs(directory, exist_ok=True)

    counter = 1
    unique_filename = f"{base_filename}-001{ext}"

    while os.path.exists(os.path.join(directory, unique_filename)):
        counter += 1
        unique_filename = f"{base_filename}-{counter:03}{ext}"

    return os.path.join(directory, unique_filename), unique_filename


def fetch_data(input_file, filename, docname):
    try:
        #Track the time of the process
        start_time = time.time()

        # Open the input and output files
        output_file, unique_filename = get_unique_filename(filename)
        with open(input_file, mode='r') as infile, open(output_file, mode='w', newline='') as outfile:
            reader = csv.reader(infile)
            writer = csv.writer(outfile)

            next(reader, None)
            
            # Loop through each row in the input file
            if filename == "sales_invoice_status":
                import_sales_invoice(reader, writer)
            elif filename == "credit_note_status":
                import_credit_note(reader, writer)
            elif filename == "debit_note_status":
                import_debit_note(reader, writer)
            elif filename == "payment_entry_status":
                import_payment_entry(reader, writer)

            end_time = time.time()
            elapsed_time = end_time - start_time
            minutes = elapsed_time // 60
            seconds = elapsed_time % 60

            frappe.msgprint(f"Process took {int(minutes)} minutes and {int(seconds):.2f} seconds.")
    
        frappe.get_doc({
            "doctype": "File",
            "file_url": f"/private/files/{unique_filename}",
            "file_name": unique_filename,
            "attached_to_doctype": "Import Data",
            "attached_to_name": docname,
            "is_private": 0
        }).insert()


        if docname:
            doc = frappe.get_doc("Import Data", docname)
            file = frappe.get_doc("File", {"file_name": unique_filename})
            doc.error_row = file.file_url
            doc.status = "Completed"
            doc.docstatus = 1
            doc.save()

    except FileNotFoundError:
        print(f"The file '{input_file}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

@frappe.whitelist(allow_guest=True)
def import_data(document_type, file , docname=None):
    file_path = frappe.get_site_path(file.strip("/"))
    
    if document_type == "Sales Invoice":
        filename = "sales_invoice_status"
        fetch_data(file_path ,filename, docname)
    elif document_type == "Credit Note":
        filename = "credit_note_status"
        fetch_data(file_path, filename, docname)
    elif document_type == "Debit Note":
        filename = "debit_note_status"
        fetch_data(file_path, filename, docname)
    elif document_type == "Payment Entry":
        filename = "payment_entry_status"
        fetch_data(file_path, filename, docname)



import frappe
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import csv
from io import StringIO

@frappe.whitelist()
def generate_excel_template(document_type, file_format = "csv"):
    # if file_format == "xlsx":
    #     output = BytesIO()
    #     wb = Workbook()
    #     ws = wb.active

    #     if(document_type == "Sales Invoice"):
    #         ws['A1'] = "ID"
    #         ws['B1'] = "Posting Date"
    #         ws['C1'] = "Debtor Code"
    #         ws['D1'] = "Reference No"
    #         ws['E1'] = "Payment Due Date"
    #         ws['F1'] = "Subtotal"
    #         ws['G1'] = "Tax"
    #         ws['H1'] = "Income Account"
    #         ws['I1'] = "Lhdn Tax"
    #         ws['J1'] = "Description"
    #         ws['K1'] = "Agent"
    #         ws['L1'] = "Tax Rate"
    #         ws['M1'] = "Project No"
    #     elif(document_type == "Credit Note"):
    #         ws['A1'] = "ID"
    #         ws['B1'] = "Posting Date"
    #         ws['C1'] = "Debtor Code"
    #         ws['D1'] = "User Remark"
    #         ws['E1'] = "Agent"
    #         ws['F1'] = "Net Total"
    #         ws['G1'] = "Total Tax"
    #         ws['H1'] = "Debit Account"
    #         ws['I1'] = "Credit Account"
    #         ws['J1'] = "Tax Account"
    #         ws['K1'] = "Description"
    #         ws['L1'] = "Reference Type"
    #         ws['M1'] = "Reference Name"
    #         ws['N1'] = "Item Classification Code"
    #         ws['O1'] = "LHDN Tax Type"
    #         ws['P1'] = "Created By"
    #         ws['Q1'] = "Tax Rate"
    #     elif(document_type == "Debit Note"):
    #         ws['A1'] = "ID"
    #         ws['B1'] = "Posting Date"
    #         ws['C1'] = "Debtor Code"
    #         ws['D1'] = "User Remark"
    #         ws['E1'] = "Agent"
    #         ws['F1'] = "Net Total"
    #         ws['G1'] = "Total Tax"
    #         ws['H1'] = "Debit Account"
    #         ws['I1'] = "Credit Account"
    #         ws['J1'] = "Tax Account"
    #         ws['K1'] = "Description"
    #         ws['L1'] = "Reference Type"
    #         ws['M1'] = "Reference Name"
    #         ws['N1'] = "Item Classification Code"
    #         ws['O1'] = "LHDN Tax Type"
    #         ws['P1'] = "Created By"
    #         ws['Q1'] = "Tax Rate"
    #     elif(document_type == "Payment Entry"):
    #         ws['A1'] = "ID"
    #         ws['B1'] = "Posting Date"
    #         ws['C1'] = "Debtor Code"
    #         ws['D1'] = "Payment Method"
    #         ws['E1'] = "Account Paid From"
    #         ws['F1'] = "Account Paid To"
    #         ws['G1'] = "Total Amount"
    #         ws['H1'] = "Remark"
    #         ws['I1'] = "Reference Type"
    #         ws['J1'] = "Reference Name"
    #         ws['K1'] = "Allocated Amount"
    #         ws['L1'] = "Cheque No"
    #         ws['M1'] = "Reference Date"

    #     wb.save(output)
    #     output.seek(0)

    #     # Send the file to browser for download
    #     frappe.response['type'] = 'binary'
    #     frappe.response['filename'] = 'Template.xlsx'
    #     frappe.response['filecontent'] = output.getvalue()
    #     frappe.response['filetype'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # elif(file_format=="csv"):

    output = StringIO()
    writer = csv.writer(output)
    document = ''
    if(document_type == "Sales Invoice"):
        document = "SalesInvoice"
        writer.writerow(["ID", "Posting Date", "Debtor Code", "Reference No", "Payment Due Date", "Subtotal", "Tax", "Income Account", "Lhdn Tax", "Description", "Agent", "Tax Rate", "Project No"])
    elif(document_type == "Credit Note"):
        document = "CreditNote"
        writer.writerow(["ID", "Posting Date", "Debtor Code", "User Remark", "Agent", "Net Total", "Total Tax", "Debit Account", "Credit Account", "Tax Account", "Description", "Reference Type", "Reference Name", "Item Classification Code", "LHDN Tax Type",  "Created By" ,"Tax Rate"])
    elif(document_type == "Debit Note"):
        document = "DebitNote"
        writer.writerow(["ID", "Posting Date", "Debtor Code",  "User Remark","Agent","Net Total","Total Tax","Debit Account","Credit Account","Tax Account","Description","Reference Type","Reference Name","Item Classification Code","LHDN Tax Type","Created By" ,"Tax Rate"])
    elif(document_type == "Payment Entry"):
        document = "PaymentEntry"
        writer.writerow(["ID","Posting Date","Debtor Code","Payment Method","Account Paid From","Account Paid To","Total Amount","Remark","Reference Type","Reference Name","Allocated Amount","Cheque No","Reference Date"])
    
    # Encode with UTF-8 BOM for Excel compatibility
    content_with_bom = '\ufeff' + output.getvalue()

    frappe.response['type'] = 'binary'
    frappe.response['filename'] = document+'Template.csv'
    frappe.response['filecontent'] = content_with_bom
    frappe.response['filetype'] = 'text/csv'